
#!/usr/bin/env python3
"""
3D Coach with wheels (light blue coach), track, labels; toggle between Manual and Excel-driven rotations.
- Uses openpyxl (no pandas) to read Excel.
- Fresh redraw each step (cleans labels, markers, texts, meshes).
- Text boxes always reflect applied values, formatted to 2 decimals.
- GPS alignment: computes ENU offsets between base (track GPS) and final (coach AID GPS),
  translates coach, and annotates both GPS points + distance gap.
Coordinate frame: X=East, Y=North, Z=Up (meters).
""" 

import math
import argparse
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d.art3d import Poly3DCollection
from matplotlib.widgets import TextBox, Button
from openpyxl import load_workbook

# Reduce title font size globally
mpl.rcParams['axes.titlesize'] = 10

# ---------------- Rotation matrices ----------------
def Rx(deg):
    a = np.deg2rad(deg); c, s = np.cos(a), np.sin(a)
    return np.array([[1, 0, 0],[0, c, -s],[0, s, c]], dtype=float)

def Ry(deg):
    a = np.deg2rad(deg); c, s = np.cos(a), np.sin(a)
    return np.array([[ c, 0, s],[ 0, 1, 0],[-s, 0, c]], dtype=float)

def Rz(deg):
    a = np.deg2rad(deg); c, s = np.cos(a), np.sin(a)
    return np.array([[c, -s, 0],[s,  c, 0],[0,  0, 1]], dtype=float)

def compose_rpy(yaw, pitch, roll, order="zyx"):
    mats = {'x': Rx(roll), 'y': Ry(pitch), 'z': Rz(yaw)}
    R = np.eye(3)
    for axis in order:
        R = mats[axis] @ R
    return R

# ---------------- Primitives ----------------
def make_box(length=1.0, width=1.0, height=1.0, center=(0,0,0)):
    L, W, H = length/2.0, width/2.0, height/2.0
    cx, cy, cz = center
    V = np.array([
        [cx - L, cy - W, cz - H],
        [cx + L, cy - W, cz - H],
        [cx + L, cy + W, cz - H],
        [cx - L, cy + W, cz - H],
        [cx - L, cy - W, cz + H],
        [cx + L, cy - W, cz + H],
        [cx + L, cy + W, cz + H],
        [cx - L, cy + W, cz + H],
    ], dtype=float)
    F = [
        [0,1,2,3],  # bottom (z=-H)
        [4,5,6,7],  # top (z=+H)
        [0,1,5,4],  # -Y side
        [1,2,6,5],  # +X side (FRONT)
        [2,3,7,6],  # +Y side
        [3,0,4,7],  # -X side (REAR)
    ]
    return V, F


def make_cylinder_y(center=(0,0,0), radius=0.5, length=0.2, n=28):
    cx, cy, cz = center
    y0, y1 = cy - length/2.0, cy + length/2.0
    thetas = np.linspace(0, 2*np.pi, num=n, endpoint=False)
    ring0, ring1 = [], []
    for th in thetas:
        x = cx + radius*np.cos(th)
        z = cz + radius*np.sin(th)
        ring0.append([x, y0, z])
        ring1.append([x, y1, z])
    V = np.array(ring0 + ring1, dtype=float)
    F = []
    for i in range(n):
        j = (i + 1) % n
        F.append([i, j, j + n, i + n])
    F.append(list(range(0, n)))                 # one cap
    F.append(list(range(2*n-1, n-1, -1)))       # other cap
    return V, F

# ---------------- Transform ----------------
def apply_rotation(V, R, origin=(0,0,0)):
    O = np.array(origin, dtype=float)
    return (R @ (V - O).T).T + O

# ---------------- Plot helpers ----------------
def plot_faces(ax, V, face_indices_list, facecolor, edgecolor='k', alpha=1.0, lw=0.7):
    polys = [[V[i] for i in f] for f in face_indices_list]
    coll = Poly3DCollection(polys, facecolors=facecolor, edgecolor=edgecolor, linewidths=lw, alpha=alpha)
    ax.add_collection3d(coll)
    return coll


def plot_mesh(ax, V, F, color='#2b6cb0', alpha=1.0, edgecolor='k', lw=0.6):
    polys = [[V[i] for i in face] for face in F]
    coll = Poly3DCollection(polys, facecolors=color, edgecolor=edgecolor, linewidths=lw, alpha=alpha)
    ax.add_collection3d(coll)
    return coll


def remove_all_artists(ax):
    """Remove all plot elements from the 3D axes, including texts.
    Ensures every redraw (Apply/Next row) starts clean."""
    for coll in list(getattr(ax, 'collections', [])):
        try: coll.remove()
        except Exception: pass
    for ln in list(getattr(ax, 'lines', [])):
        try: ln.remove()
        except Exception: pass
    for p in list(getattr(ax, 'patches', [])):
        try: p.remove()
        except Exception: pass
    for a in list(getattr(ax, 'artists', [])):
        try: a.remove()
        except Exception: pass
    for t in list(getattr(ax, 'texts', [])):
        try: t.remove()
        except Exception: pass


def set_axes_equal(ax):
    x_limits = ax.get_xlim3d(); y_limits = ax.get_ylim3d(); z_limits = ax.get_zlim3d()
    ranges = [abs(x_limits[1]-x_limits[0]), abs(y_limits[1]-y_limits[0]), abs(z_limits[1]-z_limits[0])]
    max_range = max(ranges)
    x_mid = np.mean(x_limits); y_mid = np.mean(y_limits); z_mid = np.mean(z_limits)
    ax.set_xlim3d([x_mid - max_range/2, x_mid + max_range/2])
    ax.set_ylim3d([y_mid - max_range/2, y_mid + max_range/2])
    ax.set_zlim3d([z_mid - max_range/2, z_mid + max_range/2])

# ---------------- Track (optional) ----------------
def make_track(track_length, rail_top_z,
               gauge=1.676, rail_width=0.07, rail_height=0.16,
               sleeper_len_x=0.25, sleeper_width_y=None,
               sleeper_thickness=0.12, sleeper_spacing=0.6):
    if sleeper_width_y is None:
        sleeper_width_y = gauge + 0.6
    y_rail = gauge / 2.0
    z_rail_center    = rail_top_z - rail_height/2.0
    z_sleeper_center = z_rail_center - (rail_height/2.0) - (sleeper_thickness/2.0)
    solids = []
    for sign in (-1, 1):
        V, F = make_box(length=track_length, width=rail_width, height=rail_height,
                        center=(0.0, sign*y_rail, z_rail_center))
        solids.append((V, F, {'color':'#505050','alpha':1.0,'edgecolor':'#222222','lw':0.45}))
    n_pos = int(np.floor((track_length/2.0) / sleeper_spacing))
    x_positions = np.linspace(-n_pos*sleeper_spacing, n_pos*sleeper_spacing, num=2*n_pos+1)
    for x in x_positions:
        V, F = make_box(length=sleeper_len_x, width=sleeper_width_y, height=sleeper_thickness,
                        center=(x, 0.0, z_sleeper_center))
        solids.append((V, F, {'color':'#b58840','alpha':0.98,'edgecolor':'#7a4f1d','lw':0.45}))
    return solids

# ---------------- Wheels under the coach ----------------
def make_wheels_under_coach(coach_length, gauge,
                            wheel_radius=0.46, wheel_width=0.09,
                            axle_radius=0.05, axle_spacing=2.0,
                            bogie_center_offset=None,
                            coach_bottom_z=-2.0, clearance=0.10):
    solids = []
    if bogie_center_offset is None:
        bogie_center_offset = 0.36 * coach_length
    wheel_center_z = coach_bottom_z - wheel_radius - clearance
    axle_len_y = gauge
    for bx in (-bogie_center_offset, bogie_center_offset):
        axle_xs = (bx - axle_spacing/2.0, bx + axle_spacing/2.0)
        for axx in axle_xs:
            V_ax, F_ax = make_cylinder_y(center=(axx, 0.0, wheel_center_z), radius=axle_radius, length=axle_len_y, n=24)
            solids.append((V_ax, F_ax, {'color':'#888888','alpha':1.0,'edgecolor':'#333333','lw':0.35}))
            for sign in (-1, 1):
                wy = sign * (gauge/2.0)
                V_wh, F_wh = make_cylinder_y(center=(axx, wy, wheel_center_z), radius=wheel_radius, length=wheel_width, n=28)
                solids.append((V_wh, F_wh, {'color':'#2f2f2f','alpha':1.0,'edgecolor':'#1a1a1a','lw':0.35}))
    return solids

# ---------------- GPS helpers ----------------
def haversine_distance_and_bearing(lat1, lon1, lat2, lon2):
    R = 6371000.0
    phi1 = math.radians(lat1); phi2 = math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlmb = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlmb/2)**2
    c = 2*math.atan2(math.sqrt(a), math.sqrt(1 - a))
    d = R * c
    y = math.sin(dlmb) * math.cos(phi2)
    x = math.cos(phi1)*math.sin(phi2) - math.sin(phi1)*math.cos(phi2)*math.cos(dlmb)
    theta = math.atan2(y, x)  # radians from North, clockwise
    return d, theta


def enu_offset(base_lat, base_lon, final_lat, final_lon):
    d, bearing = haversine_distance_and_bearing(base_lat, base_lon, final_lat, final_lon)
    dx_east  = d * math.sin(bearing)
    dy_north = d * math.cos(bearing)
    return dx_east, dy_north, d

# ---------------- Scene drawing ----------------
def draw_scene(ax, args, Rc, Rt, raw_coach, F_coach, raw_wheels, raw_track,
               T_coach=(0.0,0.0,0.0), T_track=(0.0,0.0,0.0), gps=None):
    remove_all_artists(ax)

    # Rotate geometries
    V_coach_rot = apply_rotation(raw_coach, Rc)
    V_coach     = V_coach_rot + np.array(T_coach)

    # Coach to draw
    plot_faces(ax, V_coach, F_coach, facecolor=args.coach_color, edgecolor='#666666', alpha=1.0, lw=0.9) # 666666grey

    # Roof center lines
    H = args.height / 2.0
    L = args.length / 2.0
    W = args.width  / 2.0
    width_line_local  = np.array([[0.0, -W, H],[0.0, +W, H]])
    length_line_local = np.array([[-L, 0.0, H],[+L, 0.0, H]])
    width_line_world  = apply_rotation(width_line_local,  Rc) + np.array(T_coach)
    length_line_world = apply_rotation(length_line_local, Rc) + np.array(T_coach)

    roof_center_local = np.array([[0.0, 0.0, H]])
    roof_center_world = apply_rotation(roof_center_local, Rc)[0] + np.array(T_coach)

    ax.scatter([roof_center_world[0]],[roof_center_world[1]],[roof_center_world[2]], color='red', s=10, depthshade=False, zorder=36)
    ax.text3D(roof_center_world[0]+1.25, roof_center_world[1]+1.25, roof_center_world[2]+1.15,
              'AID', color='crimson', fontsize=4, zorder=41)

    # Front/Rear labels (SDD) based on vertex indices, translated by T_coach
    V_vertices = apply_rotation(raw_coach, Rc) + np.array(T_coach)
    edges = [(0,1),(1,2),(2,3),(3,0),(4,5),(5,6),(6,7),(7,4),(0,4),(1,5),(2,6),(3,7)]
    mids = []
    for i,j in edges:
        mids.append(0.5*(V_vertices[i]+V_vertices[j]))
    mids = np.array(mids)

    idx_sdd = {"SDD1":5, "SDD2":6, "SDD3":2, "SDD4":1}
    front_offsets = {
        'SDD1': ( 0.30, 0.30,  0.20),
        'SDD2': ( 0.30,-0.40,  0.20),
        'SDD3': ( 0.30,-0.40, -0.30),
        'SDD4': ( 0.30, 0.30, -0.30),
    }
    for label, vidx in idx_sdd.items():
        x,y,z = V_vertices[vidx]
        dx,dy,dz = front_offsets[label]
        ax.text3D(x+dx, y+dy, z+dz, label, color='navy', fontsize=4, zorder=40)
        ax.plot([x, x+dx],[y, y+dy],[z, z+dz], color='navy', linewidth=0.8, alpha=0.7, zorder=39)

    idx_sdd_rear = {"SDD5":4, "SDD6":7, "SDD7":3, "SDD8":0}
    rear_offsets = {
        'SDD5': (-0.30, 0.30,  0.20),
        'SDD6': (-0.30,-0.40,  0.20),
        'SDD7': (-0.30,-0.40, -0.30),
        'SDD8': (-0.30, 0.30, -0.30),
    }
    for label, vidx in idx_sdd_rear.items():
        x,y,z = V_vertices[vidx]
        dx,dy,dz = rear_offsets[label]
        ax.text3D(x+dx, y+dy, z+dz, label, color='darkgreen', fontsize=4, zorder=40)
        ax.plot([x, x+dx],[y, y+dy],[z, z+dz], color='darkgreen', linewidth=0.8, alpha=0.7, zorder=39)

    all_pts = [V_coach, V_vertices, mids, width_line_world, length_line_world, roof_center_world.reshape(1,3)]

    # Wheels (follow coach translation)
    for (Vw0, Fw, st) in raw_wheels:
        Vw = apply_rotation(Vw0, Rc) + np.array(T_coach)
        st2 = {**st}
        color = st2.get('color', '#2f2f2f')
        if color in ('#2f2f2f', '#1a1a1a'):
            st2['color'] = args.wheel_color
        plot_mesh(ax, Vw, Fw, **st2)
        all_pts.append(Vw)

    # Track (own translation if needed)
    for (Vt0, Ft, st) in raw_track:
        Vt = apply_rotation(Vt0, Rt) + np.array(T_track)
        plot_mesh(ax, Vt, Ft, **st)
        all_pts.append(Vt)

    # GPS markers
    if gps is not None:
        bx, by = gps.get("base_xy", (0.0, 0.0))
        cx, cy = gps.get("coach_xy", (0.0, 0.0))
        gz     = gps.get("ground_z", 0.0)
        gap    = gps.get("distance_m", 0.0)
        ax.scatter([bx],[by],[gz], color="limegreen", s=24, depthshade=False, zorder=60)
        ax.text3D(bx, by, gz+0.2, "Track GPS", color="green", fontsize=5, zorder=61)
        ax.scatter([cx],[cy],[gz], color="red", s=24, depthshade=False, zorder=60)
        ax.text3D(cx, cy, gz+0.2, f"Coach GPS ({gap:.2f} m)", color="crimson", fontsize=5, zorder=61)
        ax.plot([bx, cx],[by, cy],[gz, gz], color="purple", linewidth=1.2, alpha=0.85, zorder=62)

    # View & limits
    ax.view_init(elev=18, azim=-55)
    all_pts = np.vstack(all_pts)
    pad = max(args.length, args.width) * 0.1
    mins = all_pts.min(axis=0) - pad
    maxs = all_pts.max(axis=0) + pad
    zoom_factor = 0.4
    centers = 0.5 * (mins + maxs)
    ranges  = (maxs - mins) * zoom_factor
    mins = centers - 0.5 * ranges
    maxs = centers + 0.5 * ranges
    ax.set_xlim(mins[0], maxs[0])
    ax.set_ylim(mins[1], maxs[1])
    ax.set_zlim(mins[2], maxs[2])
    set_axes_equal(ax)

# ---------------- Main ----------------
def main():
    ap = argparse.ArgumentParser(description="Coach + wheels + track; grey coach; labels; yaw/pitch/roll; Manual/Excel UI; GPS alignment")
    # Coach size
    ap.add_argument("--length", type=float, default=23.5, help="Coach length (X)")
    ap.add_argument("--width",  type=float, default=3.0,  help="Coach width (Y)")
    ap.add_argument("--height", type=float, default=4.0,  help="Coach height (Z)")
    # Colors
    ap.add_argument("--coach_color", type=str, default="#FFA500", help="Coach fill color ") #(light blue)  add8e6
    ap.add_argument("--wheel_color", type=str, default="#2f2f2f", help="Wheel color (dark gray)")
    # Rotations (initial)
    ap.add_argument("--yaw",   type=float, default=0.0, help="Initial Coach Yaw (deg) about Z")
    ap.add_argument("--pitch", type=float, default=0.0, help="Initial Coach Pitch (deg) about Y")
    ap.add_argument("--roll",  type=float, default=0.0, help="Initial Coach Roll (deg) about X")
    ap.add_argument("--order", type=str,   default="zyx", help="Rotation order, e.g., 'zyx' or 'xyz'")
    # Animation
    ap.add_argument("--animate", type=int, default=0, help="1 to animate (UI hidden)")
    # Track
    ap.add_argument("--show_track",  type=int,   default=1,     help="1 to draw track under the coach")
    ap.add_argument("--gauge",        type=float, default=1.676, help="Track gauge (m). 1.676 Broad, 1.435 Standard")
    ap.add_argument("--rail_width",   type=float, default=0.03,  help="Rail head width")
    ap.add_argument("--rail_height",  type=float, default=0.10,  help="Rail height")
    ap.add_argument("--sleeper_len_x",type=float, default=0.25,  help="Sleeper thickness along X")
    ap.add_argument("--sleeper_width",type=float, default=-1.0,  help="Sleeper span across Y; <=0 -> (gauge+0.6)")
    ap.add_argument("--sleeper_thk",  type=float, default=0.12,  help="Sleeper thickness (Z)")
    ap.add_argument("--sleeper_spacing", type=float, default=0.6,help="Sleeper spacing along X")
    # Wheels
    ap.add_argument("--wheel_radius", type=float, default=0.46,  help="Wheel radius")
    ap.add_argument("--wheel_width",  type=float, default=0.09,  help="Wheel thickness (Y)")
    ap.add_argument("--axle_radius",  type=float, default=0.05,  help="Axle radius")
    ap.add_argument("--axle_spacing", type=float, default=2.0,   help="Axle spacing within a bogie (X)")
    ap.add_argument("--bogie_offset", type=float, default=-1.0,  help="Bogie center offset from coach center (X). <=0 -> auto")
    ap.add_argument("--wheel_clearance", type=float, default=0.10, help="Vertical gap between coach bottom and wheel top")
    # Figure (larger image)
    ap.add_argument("--dpi", type=int, default=160, help="Figure DPI")
    ap.add_argument("--figw", type=float, default=18.0, help="Figure width (inches)")
    ap.add_argument("--figh", type=float, default=10.0, help="Figure height (inches)")

    args = ap.parse_args()

    # Geometry
    V_coach, F_coach = make_box(args.length, args.width, args.height, center=(0,0,0))
    coach_bottom_z = -args.height/2.0
    bogie_offset = None if args.bogie_offset <= 0 else args.bogie_offset
    wheel_solids = make_wheels_under_coach(
        coach_length=args.length,
        gauge=args.gauge,
        wheel_radius=args.wheel_radius,
        wheel_width=args.wheel_width,
        axle_radius=args.axle_radius,
        axle_spacing=args.axle_spacing,
        bogie_center_offset=bogie_offset,
        coach_bottom_z=coach_bottom_z,
        clearance=args.wheel_clearance
    )
    track_solids = []
    rail_top_z = coach_bottom_z - args.wheel_radius  # top of rail assumed beneath wheel center
    if args.show_track:
        track_len = args.length * 1.6
        sleeper_width = None if args.sleeper_width <= 0 else args.sleeper_width
        track_solids = make_track(
            track_length=track_len,
            rail_top_z=rail_top_z,
            gauge=args.gauge,
            rail_width=args.rail_width,
            rail_height=args.rail_height,
            sleeper_len_x=args.sleeper_len_x,
            sleeper_width_y=sleeper_width,
            sleeper_thickness=args.sleeper_thk,
            sleeper_spacing=args.sleeper_spacing
        )

    # Figure sized to half the monitor (fallback to args if detection fails)
    _figsize = (args.figw, args.figh)
    try:
        import tkinter as _tk
        _root = _tk.Tk(); _root.withdraw()
        _sw, _sh = _root.winfo_screenwidth(), _root.winfo_screenheight()
        _root.destroy()
        _figsize = ((_sw/2)/args.dpi, (_sh/2)/args.dpi)
    except Exception:
        pass
    fig = plt.figure(figsize=_figsize, dpi=args.dpi)
    ax = fig.add_subplot(111, projection='3d')
    ax.set_title(f"Coach + Wheels + Track — (order={args.order})", fontsize=8)
    ax.grid(False)
    for axis in (ax.xaxis, ax.yaxis, ax.zaxis):
        try: axis.pane.set_visible(False)
        except Exception: pass
    ax.set_xticks([]); ax.set_yticks([]); ax.set_zticks([])
    ax.set_axis_off()

    # Cache unrotated geometry
    raw_coach = V_coach.copy()
    raw_wheels = [(V.copy(), F, st) for (V, F, st) in wheel_solids]
    raw_track  = [(V.copy(), F, st) for (V, F, st) in track_solids]

    # Initial rotations
    Rc = compose_rpy(args.yaw, args.pitch, args.roll, order=args.order)
    Rt = Rc

    # Draw initial scene
    draw_scene(ax, args, Rc, Rt, raw_coach, F_coach, raw_wheels, raw_track)

    # UI controls
    if not args.animate:
        # 6 text boxes (coach & track) — initial values trimmed to 2 decimals
        ax_cyaw   = plt.axes([0.05, 0.02, 0.05, 0.04])
        ax_cpitch = plt.axes([0.16, 0.02, 0.05, 0.04])
        ax_croll  = plt.axes([0.27, 0.02, 0.05, 0.04])
        tb_cyaw   = TextBox(ax_cyaw,   "c-Y", initial=f"{args.yaw:.2f}")
        tb_cpitch = TextBox(ax_cpitch, "c-P", initial=f"{args.pitch:.2f}")
        tb_croll  = TextBox(ax_croll,  "c-R", initial=f"{args.roll:.2f}")

        ax_tyaw   = plt.axes([0.45, 0.02, 0.05, 0.04])
        ax_tpitch = plt.axes([0.56, 0.02, 0.05, 0.04])
        ax_troll  = plt.axes([0.67, 0.02, 0.05, 0.04])
        tb_tyaw   = TextBox(ax_tyaw,   "t-Y", initial=f"{args.yaw:.2f}")
        tb_tpitch = TextBox(ax_tpitch, "t-P", initial=f"{args.pitch:.2f}")
        tb_troll  = TextBox(ax_troll,  "t-R", initial=f"{args.roll:.2f}")

        # Apply button
        ax_apply = plt.axes([0.80, 0.02, 0.14, 0.05])
        btn_apply = Button(ax_apply, "Apply")
        def on_apply(event):
            try:
                cyaw   = float(tb_cyaw.text)
                cpitch = float(tb_cpitch.text)
                croll  = float(tb_croll.text)
                tyaw   = float(tb_tyaw.text)
                tpitch = float(tb_tpitch.text)
                troll  = float(tb_troll.text)
            except ValueError:
                return
            Rc_new = compose_rpy(cyaw, cpitch, croll, order=args.order)
            Rt_new = compose_rpy(tyaw, tpitch, troll, order=args.order)
            # Refresh text boxes to 2 decimals
            tb_cyaw.set_val(f"{cyaw:.2f}");   tb_cpitch.set_val(f"{cpitch:.2f}"); tb_croll.set_val(f"{croll:.2f}")
            tb_tyaw.set_val(f"{tyaw:.2f}");   tb_tpitch.set_val(f"{tpitch:.2f}"); tb_troll.set_val(f"{troll:.2f}")
            ax.set_title(
                f"Coach + Wheels + Track — Coach(yaw={cyaw:.2f}, pitch={cpitch:.2f}, roll={croll:.2f}) "
                f"Track(yaw={tyaw:.2f}, pitch={tpitch:.2f}, roll={troll:.2f}) (order={args.order})",
                fontsize=10
            )
            draw_scene(ax, args, Rc_new, Rt_new, raw_coach, F_coach, raw_wheels, raw_track)
            plt.draw()
        btn_apply.on_clicked(on_apply)

        # Mode toggle: Manual vs Excel (for COACH only)
        ax_mode = plt.axes([0.05, 0.86, 0.24, 0.06])
        btn_mode = Button(ax_mode, "In Manual")
        src_mode = "Manual"

        # Excel config & state
        EXCEL_PATH = r"C:\Users\aro\Downloads\angle_combinations_resaved.xlsx"
        EXCEL_SHEET = "Data"
        EXCEL_COLS  = (
            "yaw_deg","pitch_deg","roll_deg",
            "base_latitude","base_longitude","final_latitude","final_longitude"
        )
        excel_rows = None
        excel_idx  = 0

        def ensure_excel_loaded():
            nonlocal excel_rows
            if excel_rows is not None:
                return True
            try:
                wb = load_workbook(EXCEL_PATH, data_only=True)
                if EXCEL_SHEET not in wb.sheetnames:
                    ax.set_title(f"Sheet '{EXCEL_SHEET}' not found in workbook", fontsize=10)
                    plt.draw(); return False
                ws = wb[EXCEL_SHEET]
                headers = [cell.value for cell in ws[1]]
                col_idx = {name: i+1 for i, name in enumerate(headers)}  # 1-based
                # Required: yaw/pitch/roll
                for col in ("yaw_deg","pitch_deg","roll_deg"):
                    if col not in col_idx:
                        ax.set_title(f"Missing column '{col}' in sheet '{EXCEL_SHEET}'. Present: {headers}", fontsize=10)
                        plt.draw(); return False
                rows = []
                for r in ws.iter_rows(min_row=2, values_only=True):
                    yaw   = r[col_idx['yaw_deg']-1]
                    pitch = r[col_idx['pitch_deg']-1]
                    roll  = r[col_idx['roll_deg']-1]
                    base_lat  = r[col_idx['base_latitude']-1] if 'base_latitude' in col_idx else None
                    base_lon  = r[col_idx['base_longitude']-1] if 'base_longitude' in col_idx else None
                    final_lat = r[col_idx['final_latitude']-1] if 'final_latitude' in col_idx else None
                    final_lon = r[col_idx['final_longitude']-1] if 'final_longitude' in col_idx else None
                    rows.append({
                        'yaw_deg': yaw, 'pitch_deg': pitch, 'roll_deg': roll,
                        'base_latitude': base_lat, 'base_longitude': base_lon,
                        'final_latitude': final_lat, 'final_longitude': final_lon
                    })
                excel_rows = rows
                return True
            except Exception as e:
                ax.set_title(f"Excel load failed (openpyxl): {e}", fontsize=10)
                plt.draw(); return False

        def toggle_mode(event):
            nonlocal src_mode, excel_idx
            if src_mode == "Manual":
                if not ensure_excel_loaded():
                    return
                src_mode = "Excel"; excel_idx = 0
                n = len(excel_rows)
                btn_mode.label.set_text("from Excel")
                ax.set_title(f"from Excel (rows={n}) — Next row applies coach rotations from Excel; track stays manual.", fontsize=10)
            else:
                src_mode = "Manual"
                btn_mode.label.set_text("In Manual")
                ax.set_title("In Manual — Use text boxes and Apply.", fontsize=10)
            plt.draw()
        btn_mode.on_clicked(toggle_mode)

        # Next row (Excel)
        ax_next = plt.axes([0.80, 0.09, 0.14, 0.05])
        btn_next = Button(ax_next, "xls next_row")
        def on_next(event):
            nonlocal excel_idx
            if src_mode != "Excel":
                return
            if not ensure_excel_loaded():
                return
            if len(excel_rows) == 0:
                ax.set_title("Excel has 0 rows.", fontsize=10); plt.draw(); return

            row = excel_rows[excel_idx]
            try:
                cyaw   = float(row["yaw_deg"])
                cpitch = float(row["pitch_deg"])
                croll  = float(row["roll_deg"])
            except Exception:
                ax.set_title(f"Row {excel_idx+1}: invalid numeric values.", fontsize=10); plt.draw(); return

            # Track values from text boxes
            try:
                tyaw   = float(tb_tyaw.text)
                tpitch = float(tb_tpitch.text)
                troll  = float(tb_troll.text)
            except ValueError:
                return

            # Refresh text boxes (2 decimals)
            tb_cyaw.set_val(f"{cyaw:.2f}"); tb_cpitch.set_val(f"{cpitch:.2f}"); tb_croll.set_val(f"{croll:.2f}")
            tb_tyaw.set_val(f"{tyaw:.2f}"); tb_tpitch.set_val(f"{tpitch:.2f}"); tb_troll.set_val(f"{troll:.2f}")

            Rc_new = compose_rpy(cyaw, cpitch, croll, order=args.order)
            Rt_new = compose_rpy(tyaw, tpitch, troll, order=args.order)

            # --- GPS alignment ---
            base_lat  = row.get("base_latitude")
            base_lon  = row.get("base_longitude")
            final_lat = row.get("final_latitude")
            final_lon = row.get("final_longitude")

            T_coach = (0.0, 0.0, 0.0)
            T_track = (0.0, 0.0, 0.0)
            gps = None
            title_suffix = ""
            if (base_lat is not None and base_lon is not None and final_lat is not None and final_lon is not None):
                try:
                    dx, dy, gap_m = enu_offset(float(base_lat), float(base_lon), float(final_lat), float(final_lon))
                    # Translate coach by ENU offsets (X=East, Y=North)
                    T_coach = (dx, dy, 0.0)
                    gps = {
                        'base_xy': (0.0, 0.0),
                        'coach_xy': (dx, dy),
                        'ground_z': rail_top_z,
                        'distance_m': gap_m
                    }
                    title_suffix = f" | GPS gap={gap_m:.2f} m"
                except Exception:
                    title_suffix = " | GPS gap: error"
            else:
                title_suffix = " | GPS: missing"

            ax.set_title(
                f"Excel row {excel_idx+1}/{len(excel_rows)} — "
                f"Coach(yaw={cyaw:.2f}, pitch={cpitch:.2f}, roll={croll:.2f}) "
                f"Track(yaw={tyaw:.2f}, pitch={tpitch:.2f}, roll={troll:.2f}) (order={args.order})" + title_suffix,
                fontsize=10
            )

            draw_scene(ax, args, Rc_new, Rt_new, raw_coach, F_coach, raw_wheels, raw_track,
                       T_coach=T_coach, T_track=T_track, gps=gps)
            plt.draw()
            excel_idx = (excel_idx + 1) % len(excel_rows)
        btn_next.on_clicked(on_next)

        # Animation
    if args.animate:
        from matplotlib.animation import FuncAnimation
        base = np.array([args.yaw, args.pitch, args.roll], dtype=float)
        def update(frame):
            off = np.array([15.0*np.sin(frame*0.05), 5.0*np.sin(frame*0.07), 3.0*np.sin(frame*0.09)])
            yaw, pitch, roll = (base + off).tolist()
            Ranim = compose_rpy(yaw, pitch, roll, order=args.order)
            draw_scene(ax, args, Ranim, Ranim, raw_coach, F_coach, raw_wheels, raw_track)
            return ax.collections
        FuncAnimation(fig, update, frames=400, interval=20, blit=False)

    fig.subplots_adjust(top=0.92, bottom=0.12, left=0.05, right=0.98)
    plt.show()

if __name__ == "__main__":
    main()
